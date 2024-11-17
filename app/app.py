from flask import Flask, render_template, url_for, redirect, request, flash,send_file, Response
from flask_mysqldb import MySQL 
from flask_login import login_user, logout_user, login_required, LoginManager 
from models.ModelUser import ModelUser
from models.entities.User import User
from openpyxl.styles import PatternFill
from random import sample
from openpyxl import Workbook
import os
from io import BytesIO
import datetime
import cv2
from pyzbar.pyzbar import decode
import numpy as np
from werkzeug.utils import secure_filename 
import barcode
from barcode.writer import ImageWriter
import base64

app=Flask(__name__)
app.config['MYSQL_HOST']='localhost'
app.config['MYSQL_USER']='root'
app.config['MYSQL_PASSWORD']=''
app.config['MYSQL_DB']='crms'
mysql=MySQL(app)
logmain=LoginManager(app)
app.secret_key="mysecretkey"
usuario_logeado=ModelUser
FOTO_PREDETERMINADA='3CMP8RAD49YTLSÑHGWBU.jpg'
FOTO_PREDETERMINADA_SEXTENCION='3CMP8RAD49YTLSÑHGWBU'
@logmain.user_loader
def cargar_usuario(id):
    return ModelUser.get_by_id(mysql,id)
@app.route('/')
def Index():
    return render_template('login.html')
@app.route('/login', methods=['POST'])
def login():
    if request.method == 'POST':
        usuario=User(0,request.form['usuario'],request.form['contraseña'])    
        usuario_logeado=ModelUser.login(mysql,usuario)
        if usuario_logeado != None:
            if usuario_logeado.contraseña:
                login_user(usuario_logeado)
                id=usuario_logeado.id
                return redirect(url_for('home',id=id)) 
            else:
                flash('error!!! contraseña incorrecta')
                return render_template('login.html')
        else:
            flash('error!!! usuario no enncontrado')    
            return render_template('login.html')
    else:
        return render_template('login.html')
@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('Index'))
@app.route('/home/<id>' , methods=['POST','GET'])
@login_required
def home(id):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT Autorizado FROM persona WHERE ID_Persona={}'.format(id))
    Autorizado=cursor.fetchone()
    cursor.execute('SELECT Cantidad FROM material')
    Cantidadm=cursor.fetchall()
    cursor.execute('SELECT * FROM material')
    Materiales=cursor.fetchall()
    if request.method=='POST':
        busqueda=request.form['busqueda'].upper()
        cursor.execute('SELECT * FROM material WHERE Nombre LIKE "%{}%"'.format(busqueda))
        Materiales=cursor.fetchall()
    if (Autorizado[0])==1:
        for cantidad in Cantidadm:
            if int(cantidad[0]) <=5:
                flash('Revisa el material, esta apunto de terminarse')
                return render_template('index.html',id=id,Materiales=Materiales,Autorizado=Autorizado[0])
    return render_template('index.html',id=id,Materiales=Materiales,Autorizado=Autorizado[0])
@app.route('/personal/<id>')
@login_required
def personal(id):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT * FROM persona')
    personal=cursor.fetchall()
    cursor.execute('SELECT Autorizado FROM persona WHERE ID_Persona={}'.format(id))
    Autorizado=cursor.fetchone()
    return render_template('personal.html',id=id,personal=personal,Autorizado=Autorizado[0])
@app.route('/agregar_persona/<id>', methods=['POST'])
@login_required
def agregar_persona(id):
    nombre=request.form['Nombre']
    apellidop=request.form['Apellidop']
    apellidom=request.form['Apellidom']
    cargo=request.form['Cargo']
    usuario=request.form['Usuario']
    autorizacion=request.form.get('Autorizado')
    if autorizacion!= None:
        autorizado=1
    else:
        autorizado=0   

    contraseña=User.generar_contraseña(request.form['Contraseña'])
    cursor=mysql.connection.cursor()
    cursor.execute('INSERT INTO persona(Usuario, Contraseña, Nombre, Apellido_p, Apellido_m, Cargo,Autorizado) VALUES (%s,%s,%s,%s,%s,%s,%s)',(usuario,contraseña,nombre,apellidop,apellidom,cargo,autorizado))
    mysql.connection.commit()
    return redirect(url_for('personal',id=id))
@app.route('/eliminar_persona/<id>,<idpersona>', methods=['POST'])
@login_required
def eliminar_persona(id,idpersona):
    cursor=mysql.connection.cursor()
    cursor.execute('DELETE FROM persona WHERE ID_Persona={}'.format(idpersona))
    mysql.connection.commit()
    return redirect(url_for('personal',id=id))
@app.route('/editar_persona/<id>,<idpersona>', methods=['POST'])
@login_required
def editar_persona(id,idpersona):
    nombre=request.form['Nombre']
    apellidop=request.form['Apellidop']
    apellidom=request.form['Apellidom']
    cargo=request.form['Cargo']
    usuario=request.form['Usuario']
    autorizacion=request.form.get('Autorizado')
    if autorizacion!= None:
        autorizado=1
    else:
        autorizado=0
    cursor=mysql.connection.cursor()
    cursor.execute('UPDATE persona SET Usuario=%s, Nombre=%s, Apellido_p=%s, Apellido_m=%s, Cargo=%s,Autorizado=%s WHERE ID_Persona=%s',(usuario,nombre,apellidop,apellidom,cargo,autorizado,idpersona))
    mysql.connection.commit()
    return redirect(url_for('personal',id=id))
@app.route('/cambiar_contraseña/<id>,<idpersona>', methods=['POST'])
@login_required
def cambiar_contraseña(id,idpersona):
    contraseña=request.form['Contraseña']
    cursor=mysql.connection.cursor()
    contraseña_hash=User.generar_contraseña(contraseña)
    cursor.execute('UPDATE persona SET contraseña=%s WHERE ID_Persona=%s',(contraseña_hash,idpersona))
    mysql.connection.commit()
    return redirect(url_for('personal',id=id))

@app.route('/surtir_material/<id>', methods=['POST','GET'])
@login_required
def surtir_material(id):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT * FROM material')
    Materiales=cursor.fetchall()
    if request.method=='POST':
        busqueda=request.form['busqueda'].upper()
        cursor.execute('SELECT * FROM material WHERE Nombre LIKE "%{}%"'.format(busqueda))
        Materiales=cursor.fetchall()
    cursor.execute('SELECT Autorizado FROM persona WHERE ID_Persona={}'.format(id))
    Autorizado=cursor.fetchone()
    return render_template('surtir_material.html',id=id,Materiales=Materiales,Autorizado=Autorizado[0])
@app.route('/resurtir_material/<id>,<idmaterial>',methods=['POST'])
@login_required
def resurtir(id,idmaterial):
    cantidad=request.form['resurtir']
    cursor=mysql.connection.cursor()
    fecha_hora_actual=datetime.datetime.now()
    cursor.execute('INSERT INTO despachar_material(ID_Material,ID_Persona,Cantidad,Hora,Accion) VALUES (%s,%s,%s,%s,%s)',(idmaterial,id,cantidad,fecha_hora_actual,'Ingreso'))
    cursor.execute('UPDATE material SET Cantidad=%s WHERE ID_Material=%s',(cantidad,idmaterial))
    mysql.connection.commit()
    return redirect(url_for('surtir_material',id=id))
@app.route('/Crear_Material/<id>', methods=['POST'])
@login_required
def Crear_Material(id):
    nombre=request.form['Nombre']
    nombre=nombre.upper()
    descripcion=request.form['Descripcion']
    descripcion=descripcion.upper()
    cantidad=request.form['Cantidad']
    
    cur=mysql.connection.cursor()
    cur.execute('SELECT Codigo FROM material ORDER BY ID_Material DESC LIMIT 1')
    codigo=cur.fetchone()
    codigoentero=int(codigo[0])+1
    if (request.files['foto']).filename !='':
        foto=request.files['foto']
        cambioNombre= recibeFoto(foto)
        cur.execute('INSERT INTO material(Nombre, Descripcion, Cantidad, Foto, Codigo) VALUES (%s,%s,%s,%s,%s) ',(nombre,descripcion,cantidad,cambioNombre,codigoentero))
        mysql.connection.commit()
        flash('Material agregado')
        return redirect(url_for('surtir_material',id=id ))
    else:
        cur.execute('INSERT INTO material(Nombre, Descripcion, Cantidad, Foto, Codigo) VALUES (%s,%s,%s,%s,%s) ',(nombre,descripcion,cantidad,FOTO_PREDETERMINADA,codigoentero))
        mysql.connection.commit()
        flash('Material agregado')
        return redirect(url_for('surtir_material',id=id ))
@app.route('/Editar_Material/<id>,<idmaterial>', methods=['POST'])
@login_required
def Editar_Material(id,idmaterial):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT Foto FROM material WHERE ID_Material={}'.format(idmaterial))
    nombreFoto=cursor.fetchall()
    nombre=request.form['Nombre']
    nombre=nombre.upper()
    descripcion=request.form['Descripcion']
    descripcion=descripcion.upper()
    cantidad=request.form['Cantidad']
    foto=request.files['foto']
    if nombreFoto[0][0]!='':
        if nombreFoto[0][0]!=FOTO_PREDETERMINADA:
            if foto.filename!='':
    
                basepath = os.path.dirname (__file__) #C:\xampp\htdocs\localhost\Crud-con-FLASK-PYTHON-y-MySQL\app
                url_File = os.path.join (basepath, 'static/img', nombreFoto[0][0])
                
                cambioNombre=recibeFoto(foto)
                cursor.execute('UPDATE material SET Nombre=%s, Descripcion=%s, Cantidad=%s, Foto=%s WHERE ID_Material=%s',(nombre,descripcion,cantidad,cambioNombre,idmaterial))
                mysql.connection.commit()
            else:
                cursor.execute('UPDATE material SET Nombre=%s, Descripcion=%s, Cantidad=%s WHERE ID_Material=%s',(nombre,descripcion,cantidad,idmaterial))
                mysql.connection.commit() 
        else: 
            if foto.filename!='':
                cambioNombre= recibeFoto(foto)
                cursor.execute('UPDATE material SET Nombre=%s, Descripcion=%s, Cantidad=%s, Foto=%s WHERE ID_Material=%s',(nombre,descripcion,cantidad,cambioNombre,idmaterial))
                mysql.connection.commit()
            else:
                cursor.execute('UPDATE material SET Nombre=%s, Descripcion=%s, Cantidad=%s WHERE ID_Material=%s',(nombre,descripcion,cantidad,idmaterial))
                mysql.connection.commit()
    else:
        
                cambioNombre= recibeFoto(foto)
                cursor.execute('UPDATE material SET Nombre=%s, Descripcion=%s, Cantidad=%s, Foto=%s WHERE ID_Material=%s',(nombre,descripcion,cantidad,cambioNombre,idmaterial))
                mysql.connection.commit()
    return redirect(url_for('surtir_material',id=id ))

@app.route('/despachar_material/<id>')
@login_required
def despachar_material(id):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT Nombre,Codigo,ID_Material FROM material')
    materiales=cursor.fetchall()
    cursor.execute('SELECT Autorizado FROM persona WHERE ID_Persona={}'.format(id))
    Autorizado=cursor.fetchone()
    cursor.execute('SELECT Nombre, Codigo FROM material')
    barcode=cursor.fetchall()
    return render_template('despachar_material.html',id=id,materiales=materiales,Autorizado=Autorizado[0],barcode=barcode)
@app.route('/escanear_codigo/<id>', methods=['POST'])
@login_required
def escanear_codigo(id):
    int(id)
    codigo=request.form['Codigo']
    cantidad=request.form['Cantidad']
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT Cantidad FROM material WHERE Codigo={}'.format(codigo))
    cantidad_inicial=cursor.fetchone()
    cantidad_total=int(cantidad_inicial[0])-int(cantidad)
    cursor.execute('SELECT ID_Material FROM material WHERE Codigo={}'.format(codigo))
    idmat=cursor.fetchone()
    fecha_hora_actual=datetime.datetime.now()
    cursor.execute('INSERT INTO despachar_material(ID_Material,ID_Persona,Cantidad,Hora,Accion) VALUES (%s,%s,%s,%s,%s)',(idmat[0],id,cantidad,fecha_hora_actual,'Egreso'))
    cursor.execute('UPDATE material SET Cantidad=%s WHERE Codigo=%s',(cantidad_total,codigo))
    mysql.connection.commit()
    return redirect(url_for('home',id=id))

@app.route('/generar_codigo_barra/<id>', methods=['POST'])
def generar_codigo_barra(id):
    codigo = request.form.get('opc')
    cursor=mysql.connection.cursor()
    print(codigo)
    cursor.execute("SELECT Nombre FROM Material WHERE Codigo={}".format(codigo)) 
    nombre=cursor.fetchone()
    print(nombre)
    ean = barcode.get('ean13', codigo, writer=ImageWriter())
    buffer = BytesIO()
    ean.write(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"{nombre[0]}-barcode.png", mimetype='image/png')

@app.route('/leer_codigo_barra/<id>', methods=['GET', 'POST'])
def leer_codigo_barra(id):
    cursor=mysql.connection.cursor()
    data = request.json['image']
    print('creo que no lo leyo')
    try:
        image_data = base64.b64decode(data.split(',')[1])
        np_arr = np.frombuffer(image_data, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if decode(frame)[0]=='':
            raise TypeError('Lectura de codigo de barras fallida')
        for barcode in decode(frame):
                data = barcode.data.decode('utf-8')
                codigo=data[0:12]
        cursor.execute('SELECT Cantidad FROM material WHERE Codigo ={}'.format(codigo))
        cantidad_inicial=cursor.fetchone()
        cantidad_final= int(cantidad_inicial[0])-1
        cursor.execute('UPDATE material SET Cantidad={} WHERE Codigo={}'.format(cantidad_final,codigo))
        flash('Codigo de barras aceptado correctamente')
        mysql.connection.commit()
        
        
    except Exception:
        flash('codigo de barras fallido')

@app.route('/historial_movimiennto/<id>', methods=['POST','GET'])
def historial_movimiennto(id):
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT Autorizado FROM persona WHERE ID_Persona={}'.format(id))
    Autorizado=cursor.fetchone()
    fechar_actual=datetime.datetime.now()
    mesActual=fechar_actual.month
    anioActual=fechar_actual.year
    cursor.execute('SELECT material.Nombre, persona.Nombre, despachar_material.Cantidad, despachar_material.Hora,despachar_material.Accion FROM `despachar_material` INNER JOIN persona ON persona.ID_Persona=despachar_material.ID_Persona INNER JOIN material ON material.ID_Material=despachar_material.ID_Material WHERE MONTH(despachar_material.Hora) = %s AND YEAR(despachar_material.Hora) = %s ORDER BY despachar_material.Hora ASC',(fechar_actual.month,fechar_actual.year))
    historial_material=cursor.fetchall()
    cursor.execute('SELECT despachar_material.Hora FROM `despachar_material` ORDER BY despachar_material.Hora DESC')
    historial_material2=cursor.fetchall()
    mes=obtener_mes(historial_material2[0][0], fechar_actual)
    mes=[fila[::-1] for fila in mes[::-1]]
    if request.method=='POST':
        opcion_seleccionada = request.form.get('opcion_seleccionada')
        if len(opcion_seleccionada)==9:
            cursor.execute('SELECT material.Nombre, persona.Nombre, despachar_material.Cantidad, despachar_material.Hora,despachar_material.Accion FROM `despachar_material` INNER JOIN persona ON persona.ID_Persona=despachar_material.ID_Persona INNER JOIN material ON material.ID_Material=despachar_material.ID_Material WHERE MONTH(despachar_material.Hora) = %s AND YEAR(despachar_material.Hora) = %s',(int(opcion_seleccionada[7]),int(opcion_seleccionada[1:5])))
            historial_material=cursor.fetchall()
            mesActual=opcion_seleccionada[7]
            anioActual=opcion_seleccionada[1:5]
        else:
            cursor.execute('SELECT material.Nombre, persona.Nombre, despachar_material.Cantidad, despachar_material.Hora,despachar_material.Accion FROM `despachar_material` INNER JOIN persona ON persona.ID_Persona=despachar_material.ID_Persona INNER JOIN material ON material.ID_Material=despachar_material.ID_Material WHERE MONTH(despachar_material.Hora) = %s AND YEAR(despachar_material.Hora) = %s',(opcion_seleccionada[7:8],opcion_seleccionada[1:5]))
            historial_material=cursor.fetchall()
            mesActual=opcion_seleccionada[7:8]
            anioActual=opcion_seleccionada[1:5]
    return render_template('historial_movimiennto.html',id=id,Autorizado=Autorizado[0],historial_material=historial_material,meses=mes,mesActual=mesActual,anioActual=anioActual)
@app.route('/material_faltante/<id>')
def material_faltante(id):
    color = PatternFill(start_color="eb0202", end_color="eb0202", fill_type="solid")
    wb=Workbook()
    ws=wb.active
    ws.title="Ver Material"
    ws.merge_cells('c6:d6')
    ws.merge_cells('e6:f6')
    ws.merge_cells('g6:h6')       
    ws['c6']='ID'
    ws['e6']='Nombre' 
    ws['g6']='Cantidad'
    punter=7
    cursor=mysql.connection.cursor()
    cursor.execute('SELECT ID_Material,Nombre,Cantidad FROM material')
    materiales=cursor.fetchall()
    for material in materiales:
        punter=punter+1
        if int(material[2])<=5:
            ws['g{}'.format(punter)].fill = color
        ws.merge_cells('c{}:d{}'.format(punter,punter))
        ws.merge_cells('e{}:f{}'.format(punter,punter))
        ws.merge_cells('g{}:h{}'.format(punter,punter))
        ws['c{}'.format(punter)]=material[0]
        ws['e{}'.format(punter)]=material[1]
        ws['g{}'.format(punter)]=material[2]
    output= BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="Ver_material.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
@app.route('/descarga_hist_material/<mes>,<anio>', methods=['POST','GET'])
def descarga_hist_material(mes,anio):
    with app.app_context():
        wb=Workbook()
        ws=wb.active
        ws.title="Historial Material"
        ws.merge_cells('c6:d6')
        ws.merge_cells('e6:f6')
        ws.merge_cells('g6:h6')
        ws.merge_cells('i6:j6')
        ws.merge_cells('k6:l6')
        ws.merge_cells('c7:j7')         
        ws['c6']='Encargado'
        ws['e6']='Material' 
        ws['g6']='Cantidad'
        ws['i6']='Fecha' 
        ws['k7']='Ingreso/Egreso'
        punter=7
        cursor=mysql.connection.cursor()
        cursor.execute('SELECT material.Nombre, persona.Nombre, despachar_material.Cantidad, despachar_material.Hora,despachar_material.Accion FROM `despachar_material` INNER JOIN persona ON persona.ID_Persona=despachar_material.ID_Persona INNER JOIN material ON material.ID_Material=despachar_material.ID_Material WHERE MONTH(despachar_material.Hora) = %s AND YEAR(despachar_material.Hora) = %s ',(mes,anio))
        historial=cursor.fetchall()
        for material in historial:
            punter=punter+1
            ws.merge_cells('c{}:d{}'.format(punter,punter))
            ws.merge_cells('e{}:f{}'.format(punter,punter))
            ws.merge_cells('g{}:h{}'.format(punter,punter))
            ws.merge_cells('i{}:j{}'.format(punter,punter))
            ws.merge_cells('k{}:l{}'.format(punter,punter))
            ws['c{}'.format(punter)]=material[1]
            ws['e{}'.format(punter)]=material[0]
            ws['g{}'.format(punter)]=material[2]
            ws['i{}'.format(punter)]=material[3]
            ws['k{}'.format(punter)]=material[4]
        output= BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"Historial-{mes}-{anio}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
def recibeFoto(file):
    basepath = os.path.dirname (__file__) #La ruta donde se encuentra el archivo actual
    filename = secure_filename(file.filename) #Nombre original del archivo
    #capturando extensión del archivo ejemplo: (.png, .jpg, .pdf ...etc)
    extension           = os.path.splitext(filename)[1]
    nuevoNombreFile     = stringAleatorio() + extension
    #print(nuevoNombreFile)
    upload_path = os.path.join (basepath, 'static/img', nuevoNombreFile) 
    file.save(upload_path)
    return nuevoNombreFile
def stringAleatorio():
    straleatorio='123456789ABCDEFGHIJKLMNÑOPQRSTUVWXYZ'
    longitud = 20
    secuencia = straleatorio.upper()
    resultado_aleatorio  = sample(secuencia, longitud)
    stringaleatorio = "".join(resultado_aleatorio)
    return stringaleatorio
def obtener_mes(fecha_inicio, fecha_fin):
    meses = []
    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    # Inicializar la fecha actual para la iteración
    fecha_actual = fecha_inicio.replace(day=1)  # Ajustar al inicio del mes
    
    while fecha_actual <= fecha_fin:
        meses.append((fecha_actual.month, fecha_actual.year))
        # Incrementar el mes
        if fecha_actual.month == 12:
            fecha_actual = fecha_actual.replace(year=fecha_actual.year + 1, month=1)
        else:
            fecha_actual = fecha_actual.replace(month=fecha_actual.month + 1)

    return meses

def stauts(error):
    flash ('No no no no, por ahi no papi!!')
    return render_template('login.html')
if __name__=='__main__':
    app.register_error_handler(401, stauts)
    app.run(port=5000 , debug=True) 